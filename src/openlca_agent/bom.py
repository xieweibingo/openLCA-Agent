from __future__ import annotations

import csv
from io import StringIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from openlca_agent.models import BomItem

REQUIRED_COLUMNS = {"name", "material", "quantity", "unit"}
OPTIONAL_COLUMNS = {"supplier", "location", "notes"}

UNIT_ALIASES = {
    "kilogram": "kg",
    "kilograms": "kg",
    "kg": "kg",
    "gram": "g",
    "grams": "g",
    "g": "g",
    "piece": "piece",
    "pieces": "piece",
    "pcs": "piece",
    "unit": "piece",
    "units": "piece",
    "liter": "l",
    "liters": "l",
    "l": "l",
}


def ingest_bom(
    source_path: str | Path | None = None,
    inline_text: str | None = None,
) -> list[BomItem]:
    if source_path is None and not inline_text:
        raise ValueError("Provide source_path or inline_text for BOM ingestion.")
    rows = _read_rows(source_path=source_path, inline_text=inline_text)
    if not rows:
        return []

    normalized = [_normalize_row(row) for row in rows]
    columns = set(normalized[0])
    missing = REQUIRED_COLUMNS - columns
    if missing:
        missing_list = ", ".join(sorted(missing))
        raise ValueError(f"BOM is missing required columns: {missing_list}")

    items = []
    for index, row in enumerate(normalized, start=2):
        if not any(str(value).strip() for value in row.values() if value is not None):
            continue
        try:
            items.append(
                BomItem(
                    name=_cell(row.get("name")),
                    material=_cell(row.get("material")),
                    quantity=float(_cell(row.get("quantity"))),
                    unit=_normalize_unit(_cell(row.get("unit"))),
                    supplier=_optional_cell(row.get("supplier")),
                    location=_optional_cell(row.get("location")),
                    notes=_optional_cell(row.get("notes")),
                )
            )
        except Exception as exc:  # noqa: BLE001 - add row context for caller.
            raise ValueError(f"Invalid BOM row {index}: {exc}") from exc
    return items


def _read_rows(source_path: str | Path | None, inline_text: str | None) -> list[dict[str, Any]]:
    if inline_text:
        return list(csv.DictReader(StringIO(inline_text.strip())))
    path = Path(source_path or "")
    if not path.exists():
        raise FileNotFoundError(str(path))
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        return [dict(zip(headers, values, strict=False)) for values in rows[1:]]
    raise ValueError(f"Unsupported BOM file type: {suffix}")


def _normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    return {str(key).strip().lower(): value for key, value in row.items() if key is not None}


def _cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _optional_cell(value: Any) -> str | None:
    text = _cell(value)
    return text or None


def _normalize_unit(value: str) -> str:
    key = value.strip().lower()
    return UNIT_ALIASES.get(key, key)
