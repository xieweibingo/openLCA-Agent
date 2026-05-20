from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook

from openlca_agent.models import (
    AgentModel,
    CalculationRun,
    Hotspot,
    ImpactResult,
    ProductModel,
)


class PcfProcessContribution(AgentModel):
    process_name: str
    value: float
    unit: str = "kg CO2 eq"
    contribution: float | None = None


class PcfDeclaration(AgentModel):
    spec_version: str = "2.0"
    created: str
    product_name: str
    product_id: str
    declared_unit: str
    mass_per_declared_unit: float | None = None
    total_pcf: float
    fossil_ghg_emissions: float | None = None
    fossil_carbon_content: float | None = None
    biogenic_carbon_content: float | None = None
    impact_method: str
    processes: list[PcfProcessContribution]
    assumptions: list[str]
    missing_data: list[str]


def _climate_change_value(impacts: list[ImpactResult]) -> float | None:
    for impact in impacts:
        key = impact.impact_category.lower()
        if "climate change" in key or "global warming" in key or "gwp" in key:
            return impact.value
    return None


def _build_processes(
    hotspots: list[Hotspot],
) -> list[PcfProcessContribution]:
    return [
        PcfProcessContribution(
            process_name=h.name,
            value=h.value,
            unit=h.unit or "kg CO2 eq",
            contribution=h.contribution,
        )
        for h in hotspots
    ]


def _assumptions_list(
    run: CalculationRun,
    product_model: ProductModel | None,
) -> list[str]:
    items = list(run.assumptions)
    if product_model:
        items.extend(product_model.assumptions)
    return items


def _missing_data_list(
    run: CalculationRun,
    product_model: ProductModel | None,
) -> list[str]:
    items = list(run.missing_data)
    if product_model:
        items.extend(product_model.missing_data)
    return list(dict.fromkeys(items))


def build_pcf(
    run: CalculationRun,
    product_model: ProductModel | None = None,
) -> PcfDeclaration:
    total_pcf = _climate_change_value(run.total_impacts) or 0.0
    product_name = (
        product_model.product_name
        if product_model
        else run.product_system.get("name", "Unknown")
    )
    product_id = run.product_system.get("id", "")
    declared_unit = (
        product_model.functional_unit if product_model else "1 piece"
    )
    impact_method = run.impact_method.get("name", "unknown")

    return PcfDeclaration(
        created=datetime.now(UTC).isoformat(),
        product_name=product_name,
        product_id=product_id,
        declared_unit=declared_unit,
        total_pcf=total_pcf,
        fossil_ghg_emissions=total_pcf,
        impact_method=impact_method,
        processes=_build_processes(run.hotspots),
        assumptions=_assumptions_list(run, product_model),
        missing_data=_missing_data_list(run, product_model),
    )


def export_pcf_json(pcf: PcfDeclaration, path: Path) -> str:
    path.write_text(pcf.model_dump_json(indent=2), encoding="utf-8")
    return str(path)


def export_pcf_xlsx(pcf: PcfDeclaration, path: Path) -> str:
    workbook = Workbook()

    meta = workbook.active
    meta.title = "PCF_INFO"
    _append_rows(
        meta,
        [
            ["FIELD", "VALUE"],
            ["spec_version", pcf.spec_version],
            ["created", pcf.created],
            ["product_name", pcf.product_name],
            ["product_id", pcf.product_id],
            ["declared_unit", pcf.declared_unit],
            ["total_pcf (kg CO2 eq)", str(pcf.total_pcf)],
            ["fossil_ghg_emissions", str(pcf.fossil_ghg_emissions or "")],
            ["fossil_carbon_content", str(pcf.fossil_carbon_content or "")],
            ["biogenic_carbon_content", str(pcf.biogenic_carbon_content or "")],
            ["impact_method", pcf.impact_method],
        ],
    )

    if pcf.processes:
        proc = workbook.create_sheet("PROCESSES")
        _append_rows(
            proc,
            [
                ["PROCESS_NAME", "VALUE", "UNIT", "CONTRIBUTION"],
                *[
                    [
                        pp.process_name,
                        pp.value,
                        pp.unit,
                        pp.contribution if pp.contribution is not None else "",
                    ]
                    for pp in pcf.processes
                ],
            ],
        )

    if pcf.assumptions:
        asm = workbook.create_sheet("ASSUMPTIONS")
        _append_rows(asm, [["ASSUMPTION"]] + [[a] for a in pcf.assumptions])

    if pcf.missing_data:
        mis = workbook.create_sheet("MISSING_DATA")
        _append_rows(mis, [["MISSING_DATA"]] + [[m] for m in pcf.missing_data])

    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(
                max_length + 2, 80
            )

    workbook.save(path)
    return str(path)


def _append_rows(sheet, rows: list[list[object]]) -> None:
    for row in rows:
        sheet.append(row)
