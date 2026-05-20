from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path

from openpyxl import Workbook

from openlca_agent.models import CalculationRun, ComplianceReport, ProductModel

DEFAULT_STANDARDS = ["PCF", "EPD", "CBAM", "DPP"]


def export_result(
    run: CalculationRun,
    output_dir: str | Path,
    formats: Iterable[str] = ("xlsx",),
) -> dict[str, str]:
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    files: dict[str, str] = {}
    requested = {item.lower() for item in formats}

    if "xlsx" in requested or "excel" in requested:
        path = output_path / "results.xlsx"
        _write_results_workbook(run, path)
        files["xlsx"] = str(path)
    if "json" in requested:
        path = output_path / "results.json"
        path.write_text(run.model_dump_json(indent=2), encoding="utf-8")
        files["json"] = str(path)
    if "csv" in requested:
        path = output_path / "total_impacts.csv"
        _write_total_impacts_csv(run, path)
        files["csv"] = str(path)

    return files


def generate_compliance_report(
    run: CalculationRun,
    product_model: ProductModel | None,
    output_dir: str | Path,
    standards: list[str] | None = None,
) -> ComplianceReport:
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    standards = standards or DEFAULT_STANDARDS
    excel_files = export_result(run, output_path, formats=["xlsx"])
    notes = {
        standard: _standard_note(standard, run, product_model)
        for standard in standards
    }
    product_system_name = run.product_system.get("name", "product system")
    method_version = run.impact_method.get("name") or run.impact_method.get("id") or "unknown"
    report = ComplianceReport(
        summary=f"draft / for review assessment for {product_system_name}",
        method_version=str(method_version),
        data_sources=_data_sources(run),
        assumptions=_assumptions(run, product_model),
        impact_results=run.total_impacts,
        hotspots=run.hotspots,
        missing_data=_missing_data(run, product_model),
        pcf_epd_cbam_dpp_notes=notes,
        files={"excel": excel_files["xlsx"]},
    )
    markdown_path = output_path / "report.md"
    markdown_path.write_text(_render_markdown(report, run, product_model), encoding="utf-8")
    report.files["markdown"] = str(markdown_path)
    return report


def _write_results_workbook(run: CalculationRun, path: Path) -> None:
    workbook = Workbook()
    meta = workbook.active
    meta.title = "META_INFO"
    _append_rows(
        meta,
        [
            ["KEY", "VALUE"],
            ["RUN_ID", run.run_id],
            ["PRODUCT_SYSTEM_NAME", run.product_system.get("name", "")],
            ["PRODUCT_SYSTEM_ID", run.product_system.get("id", "")],
            ["IMPACT_METHOD_NAME", run.impact_method.get("name", "")],
            ["IMPACT_METHOD_ID", run.impact_method.get("id", "")],
        ],
    )

    impacts = workbook.create_sheet("TOTAL_IMPACTS")
    _append_rows(
        impacts,
        [["IMPACT_CATEGORY_NAME", "VALUE", "REFERENCE_UNIT"]]
        + [
            [impact.impact_category, impact.value, impact.unit or ""]
            for impact in run.total_impacts
        ],
    )

    hotspots = workbook.create_sheet("HOTSPOTS")
    _append_rows(
        hotspots,
        [["NAME", "VALUE", "UNIT", "CONTRIBUTION", "DIMENSION"]]
        + [
            [
                hotspot.name,
                hotspot.value,
                hotspot.unit or "",
                hotspot.contribution if hotspot.contribution is not None else "",
                hotspot.dimension,
            ]
            for hotspot in run.hotspots
        ],
    )

    assumptions = workbook.create_sheet("ASSUMPTIONS")
    _append_rows(assumptions, [["ASSUMPTION"]] + [[item] for item in run.assumptions])

    missing = workbook.create_sheet("MISSING_DATA")
    _append_rows(missing, [["MISSING_DATA"]] + [[item] for item in run.missing_data])

    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 80)
    workbook.save(path)


def _append_rows(sheet, rows: list[list[object]]) -> None:
    for row in rows:
        sheet.append(row)


def _write_total_impacts_csv(run: CalculationRun, path: Path) -> None:
    lines = ["IMPACT_CATEGORY_NAME,VALUE,REFERENCE_UNIT"]
    for impact in run.total_impacts:
        lines.append(f"{impact.impact_category},{impact.value},{impact.unit or ''}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _render_markdown(
    report: ComplianceReport,
    run: CalculationRun,
    product_model: ProductModel | None,
) -> str:
    model_name = (
        product_model.product_name
        if product_model
        else run.product_system.get("name", "Unknown")
    )
    product_system = run.product_system
    impact_method = run.impact_method
    lines = [
        f"# {model_name} LCA Compliance Report",
        "",
        "> Status: draft / for review. This is not a verified third-party certification report.",
        "",
        "## Project Summary",
        report.summary,
        "",
        "## Functional Unit",
        product_model.functional_unit if product_model else "Not specified",
        "",
        "## System Boundary",
        "Foreground product model connected to available background processes "
        "in the currently opened openLCA database.",
        "",
        "## Database And Method Version",
        f"- Product system: {product_system.get('name', '')} ({product_system.get('id', '')})",
        f"- Impact method: {impact_method.get('name', '')} ({impact_method.get('id', '')})",
        "",
        "## Modeling Assumptions",
        *_bullet_list(report.assumptions),
        "",
        "## Process Mapping",
        *_mapping_table(product_model),
        "",
        "## Data Quality Indicators (DQI)",
        *_dqi_table(product_model),
        "",
        "## Total Impact Results",
        *_impact_table(run),
        "",
        "## Hotspot Analysis",
        *_hotspot_table(run),
        "",
        "## Missing Data",
        *_bullet_list(report.missing_data),
        "",
        "## PCF / EPD / CBAM / DPP Draft Conclusions",
        *_standard_lines(report),
        "",
        "## Manual Review Items",
        "- Review all process mapping decisions with confidence below 0.85.",
        "- Confirm functional unit, system boundary, allocation choices, "
        "and method version before external use.",
    ]
    return "\n".join(lines).rstrip() + "\n"


def _bullet_list(items: list[str]) -> list[str]:
    return [f"- {item}" for item in items] if items else ["- None recorded."]


def _mapping_table(product_model: ProductModel | None) -> list[str]:
    if not product_model or not product_model.mapping_decisions:
        return ["No product mapping decisions recorded."]
    lines = [
        "| BOM item | Material | Selected process | Confidence | Reason |",
        "| --- | --- | --- | --- | --- |",
    ]
    for decision in product_model.mapping_decisions:
        selected = decision.selected_candidate.name if decision.selected_candidate else "Unresolved"
        lines.append(
            f"| {decision.item.name} | {decision.item.material} | {selected} | "
            f"{decision.confidence:.2f} | {decision.reason or decision.unresolved_reason or ''} |"
        )
    return lines


def _dqi_table(product_model: ProductModel | None) -> list[str]:
    if not product_model or not product_model.mapping_decisions:
        return ["No DQI data recorded."]
    lines = [
        "| BOM item | Selected process | DQI (overall) | Band | Flags |",
        "| --- | --- | ---: | --- | --- |",
    ]
    for decision in product_model.mapping_decisions:
        selected = decision.selected_candidate.name if decision.selected_candidate else "Unresolved"
        dqi = decision.dqi
        if dqi:
            flags = "; ".join(dqi.flags[:2])
            lines.append(
                f"| {decision.item.name} | {selected} | "
                f"{dqi.overall:.0f} | {dqi.confidence_band} | {flags} |"
            )
        else:
            lines.append(
                f"| {decision.item.name} | {selected} | — | — | No DQI |"
            )
    return lines


def _impact_table(run: CalculationRun) -> list[str]:
    if not run.total_impacts:
        return ["No impact results recorded."]
    lines = ["| Impact category | Value | Unit |", "| --- | ---: | --- |"]
    for impact in run.total_impacts:
        lines.append(f"| {impact.impact_category} | {impact.value:.6g} | {impact.unit or ''} |")
    return lines


def _hotspot_table(run: CalculationRun) -> list[str]:
    if not run.hotspots:
        return ["Hotspot detail unavailable or not calculated."]
    lines = ["| Name | Value | Unit | Contribution |", "| --- | ---: | --- | ---: |"]
    for hotspot in run.hotspots:
        contribution = "" if hotspot.contribution is None else f"{hotspot.contribution:.2%}"
        lines.append(
            f"| {hotspot.name} | {hotspot.value:.6g} | "
            f"{hotspot.unit or ''} | {contribution} |"
        )
    return lines


def _standard_lines(report: ComplianceReport) -> list[str]:
    return [f"- {standard}: {note}" for standard, note in report.pcf_epd_cbam_dpp_notes.items()]


def _standard_note(
    standard: str,
    run: CalculationRun,
    product_model: ProductModel | None,
) -> str:
    missing = _missing_data(run, product_model)
    suffix = " Manual review is required before external disclosure."
    if missing:
        suffix = f" Missing data remains: {', '.join(missing)}. Manual review is required."
    return f"Draft {standard} interpretation generated from current model and LCIA results.{suffix}"


def _data_sources(run: CalculationRun) -> list[str]:
    return [
        f"Product system: {run.product_system.get('name', '')}",
        f"Impact method: {run.impact_method.get('name', '')}",
    ]


def _assumptions(run: CalculationRun, product_model: ProductModel | None) -> list[str]:
    assumptions = list(run.assumptions)
    if product_model:
        assumptions.extend(product_model.assumptions)
    return assumptions


def _missing_data(run: CalculationRun, product_model: ProductModel | None) -> list[str]:
    missing = list(run.missing_data)
    if product_model:
        missing.extend(product_model.missing_data)
    return list(dict.fromkeys(missing))
